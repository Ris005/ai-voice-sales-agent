"""Excel-backed CRM. Reads leads, filters pending ones, and writes every call
outcome back to the sheet. openpyxl keeps a single canonical .xlsx as the source
of truth so the file can be opened in Excel/Sheets at any time."""
from __future__ import annotations

import threading
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from . import config

# All writes go through one lock — call outcomes can land concurrently.
_LOCK = threading.Lock()

# Column order is the contract between the sheet and the app.
COLUMNS = [
    "Lead ID",
    "Name",
    "Phone",
    "Company",
    "Status",              # pending | completed | opted-out
    "Call Status",         # booked | not-interested | no-answer | callback | ...
    "Lead Qualification",  # hot | warm | cold | unqualified
    "Conversation Summary",
    "Customer Requirements",
    "Objections Raised",
    "Follow-up Date",
    "Meeting Date & Time",
    "Last Contacted",
]

SEED_LEADS = [
    ["L-001", "Rahul Sharma", "+91 98100 11223", "BrightRetail Pvt Ltd", "pending", "", "", "", "", "", "", "", ""],
    ["L-002", "Priya Nair", "+91 99000 44556", "Coastal Logistics", "pending", "", "", "", "", "", "", "", ""],
    ["L-003", "Aakash Verma", "+91 90000 77889", "MediCare Clinics", "pending", "", "", "", "", "", "", "", ""],
    ["L-004", "Sneha Kulkarni", "+91 98888 33221", "EduSpark Academy", "pending", "", "", "", "", "", "", "", ""],
    ["L-005", "David Menezes", "+91 97777 55443", "Menezes & Co", "completed", "booked", "hot", "Already onboarded.", "", "", "", "2026-06-01 10:00", "2026-05-28 14:12"],
    ["L-006", "Farah Khan", "+91 96666 22110", "Khan Interiors", "opted-out", "do-not-call", "", "Requested no contact.", "", "", "", "", "2026-05-20 09:30"],
    ["L-007", "Vikram Singh", "+91 95555 66778", "Singh Motors", "pending", "", "", "", "", "", "", "", ""],
    ["L-008", "Meera Iyer", "+91 94444 88990", "Iyer Textiles", "pending", "", "", "", "", "", "", "", ""],
]


def ensure_crm_exists() -> None:
    """Create a seeded workbook on first run so the app is demoable instantly."""
    path: Path = config.CRM_PATH
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(COLUMNS)
    for row in SEED_LEADS:
        ws.append(row)
    wb.save(path)


def _load():
    ensure_crm_exists()
    wb = load_workbook(config.CRM_PATH)
    return wb, wb["Leads"] if "Leads" in wb.sheetnames else wb.active


def _row_to_dict(headers: list[str], row: tuple) -> dict[str, Any]:
    record = {}
    for key, value in zip(headers, row):
        record[key] = "" if value is None else value
    return record


def get_leads() -> list[dict[str, Any]]:
    wb, ws = _load()
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) for h in rows[0]]
    leads = [_row_to_dict(headers, r) for r in rows[1:] if any(c is not None for c in r)]
    return leads


def get_lead(lead_id: str) -> dict[str, Any] | None:
    for lead in get_leads():
        if str(lead.get("Lead ID")) == str(lead_id):
            return lead
    return None


def pending_leads() -> list[dict[str, Any]]:
    """Only leads that still need a call — skips completed and opted-out."""
    return [l for l in get_leads() if str(l.get("Status", "")).lower() == "pending"]


def stats() -> dict[str, int]:
    leads = get_leads()
    def count(status: str) -> int:
        return sum(1 for l in leads if str(l.get("Status", "")).lower() == status)
    booked = sum(1 for l in leads if str(l.get("Call Status", "")).lower() == "booked")
    return {
        "total": len(leads),
        "pending": count("pending"),
        "completed": count("completed"),
        "opted_out": count("opted-out"),
        "booked": booked,
    }


def update_lead(lead_id: str, updates: dict[str, Any]) -> dict[str, Any] | None:
    """Write call outcome fields back to the row identified by Lead ID."""
    with _LOCK:
        wb, ws = _load()
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h) for h in rows[0]]
        col_index = {name: i + 1 for i, name in enumerate(headers)}

        target_row = None
        for r_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=r_idx, column=col_index["Lead ID"]).value
            if str(cell) == str(lead_id):
                target_row = r_idx
                break
        if target_row is None:
            return None

        updates = {**updates, "Last Contacted": datetime.now().strftime("%Y-%m-%d %H:%M")}
        for key, value in updates.items():
            if key in col_index:
                ws.cell(row=target_row, column=col_index[key]).value = value

        wb.save(config.CRM_PATH)
        return get_lead(lead_id)
